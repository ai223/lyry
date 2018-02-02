'''
Created on Jan 13, 2018

@author: aleks
'''

import sys

from bs4 import BeautifulSoup
import requests
import datetime

import csv
import openpyxl
from openpyxl.styles.fonts import Font


# global variables
chart_names = []
charts_dict = {}

# string constants
sep = "\n" + "="*105 + "\n"
msg1 = sep + "Welcome to Lyry, a program which allows you to obtain song lyrics from any billboard chart."


"""  """
def script_start(msg) :
    # obtain all billboard.com chart names and links 
    charts_tuples = pull_chart_names()
    
    global chart_names
    chart_names = charts_tuples[0]
    global charts_dict
    charts_dict = charts_tuples[1]
    
    # begin user interaction
    print(msg)
    display_options()
    chart_link = get_chart_selection()
    chart_url_list = get_date_selection(chart_link)
    all_songs_and_artists_dict = get_song_and_artist_names(chart_url_list)
    write_billboard_lyrics_csv2(chart_url_list, all_songs_and_artists_dict)
    go_again()

""" Accesses billboard.com to get a list of all currently offered charts and their 
    respective bilbloard.com links.
    
    The function returns both a list of chart names (for pretty printing) and a 
    dictionary that returns a billboard.com link when queried with the name of a
    particular chart.

    @return name_list: a list of all billboard chart names, as appearing on 
                       https://www.billboard.com/charts 
    
    @return chart_dict: a dictionary where the key-value pairs are chart names and
                        their respective billboard.com links """
def pull_chart_names() :
    url = "https://www.billboard.com/charts"
    html = requestURL(url)
    
    soup = BeautifulSoup(html, 'lxml')
    chart_links = soup.select('.chart-row__chart-link')
    
    name_list, href_list= [], []
    for i in chart_links :
        name_list.append(i.string)
        href_list.append(i['href'])
    chart_dict = dict(zip(name_list, href_list))

    return (name_list, chart_dict)


""" Prompts the user to select a billboard chart for lyrics extraction.

    The user can make a few different choices. They can enter a number corresponding to
    a particular billboard chart (i.e. "40" representing "Hot Country Songs") or they can
    enter the chart name in full.  """
def get_chart_selection() :
    chart = input("Please enter a billboard chart name or number: ").strip()   

    if chart == "ls" :
        ls(chart_names)
        return get_chart_selection()
    elif chart == "q" :
        print(sep, "\nThanks for using the program!")
        sys.exit()
    else: 
        # if user enters chart number
        if chart.isdigit() :
            pos = int(chart) - 1
            
            if pos < 1 or pos > len(chart_names) :
                print(sep + "\nInvalid number selection. Please try again.\n")
                return get_chart_selection() 
            
            chart_name = chart_names[pos]
            earliest_date = getEarliestDate(charts_dict[chart_name])
            
            print(sep + "\nYou have selected: " + chart_name + ".")
            print("\nNow enter a date or range of dates from which " +
              "you would like to obtain lyrics.\n\n(Note: dates need to be entered in " +
              "a valid YYYY-MM-DD format. To enter a range of dates, simply enter two " +
              "dates\nseparated by a single space (e.g. '2011-01-03 2011-07-21'). Hit " +
              "'Enter' without typing anything to obtain the\ncurrent chart by default.)" + 
              "\n\nThe earliest available date for " + chart_name + " is: " + earliest_date)
                        
            return charts_dict[chart_name]
        
        # if user enters chart name
        elif all((x.isalpha() or x.isspace()) and not(x.isdigit()) for x in chart) :
            print("You have selected: " + chart + ".")
            return charts_dict[chart]
        
        # if user enters an invalid option
        else :
            print(sep + "\nYou have made an invalid selection " + 
                        "(name with number is not allowed) Try again.\n")
            return get_chart_selection()


""" Prompt user for valid date or range of dates corresponding to available 
    Billboard chart(s). """
def get_date_selection(chart_link) :
 
    date = input("Please enter your date selection: ").strip()
    
    if date == 'q' :
        print(sep, "\nThanks for using the program!")
        sys.exit()
    
    elif date == '' :  
        today = datetime.date.today()
        # move date up to Saturday, if not on Saturday
        date = resetDate(today)
        
        return ["https://www.billboard.com/" + chart_link + "/" + str(date)]
    
    elif len(date) == 10 :
        date = validate(date)
        if date == '#' :
            return get_date_selection(chart_link) 
        
        return ["https://www.billboard.com/" + chart_link + "/" + str(date)]
    
    elif len(date) == 21 :
        split = date.find(" ")
        date1 = date[:split]
        date2 = date[split+1:]

        all_dates = validate(date1, date2)
        
        func = lambda each_date : "https://www.billboard.com/" + chart_link + "/" + str(each_date)
        all_chart_urls = list( map(func, all_dates) )

        return all_chart_urls
    else :
        print("Invalid date selection: " + date + ". Please try again or enter 'q' to stop." )
        return get_date_selection(chart_link)


""" """
def get_song_and_artist_names(chart_url_list) :
    all_songs_and_artists_dict = {}
    
    for url in chart_url_list :
        
        html = requestURL(url)
        soup = BeautifulSoup(html, 'lxml')
        song_tags = soup.select(".chart-row__song")
        artist_tags = soup.select(".chart-row__artist")
        position_tags = soup.select(".chart-row__current-week") 
        last_week_tags = soup.select(".chart-row__last-week")
        weeks_on_tags = soup.select(".chart-row__weeks-on-chart")
        assert len(song_tags) == len(artist_tags)

        song_list = []
        [song_list.append(tag.string) for tag in song_tags]
        
        # [1:-1] slice removes bookend newline characters from the artist tag
        artist_list = []
        [artist_list.append(tag.string[1:-1]) for tag in artist_tags]
        
        position_list = []
        [position_list.append(tag.string) for tag in position_tags]
        
        last_week_list = []
        [last_week_list.append(tag.text) for tag in last_week_tags]
        # billboard.com duplicates the ".chart-row__last-week" tags, so remove half
        # split() converts "Last week: X" into a three-item list, and [2] chooses the number
        last_week_list = [last_week_list[x].split(" ")[2] for x in range(0, len(song_list)*2) if x%2 == 0]
        
        weeks_on_list = []
        # [13:] slice removes "Wks on chart" from the string, leaving just the number
        [weeks_on_list.append(tag.text[13:]) for tag in weeks_on_tags]
        
        song_and_artist_names = list(zip(song_list, artist_list, 
                                         position_list, last_week_list, weeks_on_list))
        
        all_songs_and_artists_dict[url] = song_and_artist_names
        
    
    return all_songs_and_artists_dict


""" """
def extract_billboard_lyrics(song_name, artist_name) :
    
    metrolyrics_url = create_metrolyrics_url(song_name, artist_name)
    lyrics = extract_metrolyrics(metrolyrics_url)
    if lyrics == '' and ' and ' in artist_name.lower() :
        pos = artist_name.lower().find(' and')
        alt_artist_name = artist_name[:pos]
        metrolyrics_url = create_metrolyrics_url(song_name, alt_artist_name)
        lyrics = extract_metrolyrics(metrolyrics_url)
                    
    if lyrics == '' :    
        genius_url = create_genius_url(song_name, artist_name)
        lyrics = extract_genius(genius_url)
        if lyrics == '' and ' and ' in artist_name.lower() :
            pos = artist_name.lower().find(' and')
            alt_artist_name = artist_name[:pos]
            genius_url = create_genius_url(song_name, alt_artist_name)
            lyrics = extract_genius(genius_url)
        
    if lyrics == '' :
        az_url = create_azlyrics_url(song_name, artist_name)
        lyrics = extract_azlyrics(az_url)
        if lyrics == '' and ' and ' in artist_name.lower() :
            pos = artist_name.lower().find(' and')
            alt_artist_name = artist_name[:pos]
            az_url = create_azlyrics_url(song_name, alt_artist_name)
            lyrics = extract_azlyrics(az_url)

    return lyrics


""" """
def write_billboard_lyrics_txt(chart_url_list, all_songs_and_artists_dict) :
    for chart_url in chart_url_list :
        filename = createFilename(chart_url)
        
        print("Extracting and writing lyrics to file...")
        
        with open(filename, 'w', encoding="utf-8") as file: 
            
            song_and_artist_name_list = all_songs_and_artists_dict[chart_url]
            
            tot_num_songs = len(song_and_artist_name_list)
            tot_num_missd = 0
            def tot_num_found() : return tot_num_songs - tot_num_missd
            def percent_yield() : return ((tot_num_songs - tot_num_missd)/tot_num_songs)*100
            
            for song_and_artist_pair in song_and_artist_name_list :
                song_name   = song_and_artist_pair[0]
                artist_name = song_and_artist_pair[1]
                position = song_and_artist_pair[2]
                last_week = song_and_artist_pair[3]
                weeks_on = song_and_artist_pair[4]
                
                lyrics = extract_billboard_lyrics(song_name, artist_name)
                
                if lyrics == '' : 
                    tot_num_missd += 1
                    continue
                
                # write song title, lyrics, current position, last week's position and 
                file.write(artist_name + ": " + song_name)
                file.write("\nPosition: " + position + "\n")
                file.write(last_week + "\n")
                file.write("Weeks on chart: " + weeks_on + "\n\n")
                file.write(lyrics)
                file.write(sep)   
                
            file.write("Number of total songs: " + str(tot_num_songs) + "\n")
            file.write("Number of found lyrics: " + str(tot_num_found()) + "\n")
            file.write("Success rate: " + str(percent_yield()) + "%")               


""" """
def write_billboard_lyrics_csv(chart_url_list, all_songs_and_artists_dict) :
    
    for chart_url in chart_url_list :
        filename = createFilename(chart_url) + ".csv"
        
        header = ["Position", "Song", "Lyrics", "Last Week", "Weeks on Chart"]
        with open(filename, "w", encoding="utf-8") as file :
            writer = csv.writer(file)
            writer.writerow(header)
            
            song_and_artist_list = all_songs_and_artists_dict[chart_url]
            
            tot_num_songs = len(song_and_artist_list)
            tot_num_missd = 0
            def tot_num_found() : return tot_num_songs - tot_num_missd
            def percent_yield() : return ((tot_num_songs - tot_num_missd)/tot_num_songs)*100
            
            for song_and_artist_iter in song_and_artist_list :
                song_name   = song_and_artist_iter[0]
                artist_name = song_and_artist_iter[1]
                position = song_and_artist_iter[2]
                last_week = song_and_artist_iter[3]
                weeks_on = song_and_artist_iter[4]
                
                lyrics = extract_billboard_lyrics(song_name, artist_name)
                
                if lyrics == '' : 
                    tot_num_missd += 1
                    continue
                                
                row = [position, artist_name, song_name, lyrics, last_week, weeks_on]
                writer.writerow(row)  
            
            success_row_header = ["Total Songs", "Found Songs", "Success Rate"]
            success_row_report = [str(tot_num_songs), str(tot_num_found()), str(percent_yield())]   
            writer.writerow(success_row_header)
            writer.writerow(success_row_report)
    
    return


"""  """
def write_billboard_lyrics_csv2(chart_url_list, all_songs_and_artists_dict) :
    wb = openpyxl.Workbook()

    # create filename for the .csv file or get it from user 
    default_filename = createFilename(chart_url_list[0], csv=True)[0]
    
    print(sep + "\nPlease choose a filename to save your data. Hitting 'Enter' will " + 
          "select the default filename: '" + default_filename + "'\n")
    filename = input("Enter your filename: ").strip()
    
    if filename == 'q' :
        print(sep, "\nThanks for using the program!")
        sys.exit()        
    elif filename == '' : 
        filename = default_filename + ".xlsx"
    else :
        filename = filename + ".xlsx"    
    wb.save(filename)
    
    for chart_url in chart_url_list :
        print(sep + "\n\tExtracting and writing lyrics from:", chart_url, "\n")
        
        # get the list of songs and artists from each chart
        song_and_artist_list = all_songs_and_artists_dict[chart_url]
        
        # track number of successfully extracted lyrics per chart
        tot_num_songs = len(song_and_artist_list)
        tot_num_missd = 0
        def tot_num_found() : return tot_num_songs - tot_num_missd
        def percent_yield() : return ((tot_num_songs - tot_num_missd)/tot_num_songs)*100
        
        # create new worksheet for each chart
        sheetname = createFilename(chart_url, csv=True)[1]
        wb.create_sheet(sheetname)
        ws = wb.get_sheet_by_name(sheetname)
        
        # set column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 60
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10
        
        header = ['Position', 'Song Name', 'Artist Name', 'Lyrics', 'Last Week', "Weeks On"]
        ws.append(header)
        for cell in ws[1:1] :
            cell.font = Font(bold = True)
        
        for song_and_artist_iter in song_and_artist_list :
            song_name   = song_and_artist_iter[0]
            artist_name = song_and_artist_iter[1]
            position = song_and_artist_iter[2]
            last_week = song_and_artist_iter[3]
            weeks_on = song_and_artist_iter[4]        
        
            # sometimes two songs are grouped together in one position
            if "/" in song_name :
                print("Get's here!")
                tot_num_songs += 1
                rows = deal_with_two_songs(filename, position, song_name, artist_name, last_week, weeks_on)
                ws.append(rows[0])
                ws.append(rows[1])
                tot_num_missd += rows[2]
                wb.save(filename)
                continue
        
            lyrics = extract_billboard_lyrics(song_name, artist_name)

            # if lyrics were not found
            if lyrics == '' : 
                tot_num_missd += 1
                row = [position, song_name, artist_name, 'COULD NOT FIND']
                ws.append(row)
                wb.save(filename)
                continue
        
            row = [position, song_name, artist_name, lyrics, last_week, weeks_on]
            ws.append(row)
            wb.save(filename)
        
        success_row_header = ["Total Songs", "Found Songs", "Success Rate"]
        success_row_report = [str(tot_num_songs), str(tot_num_found()), str(percent_yield())]   
        ws.append([''])
        ws.append(success_row_header)
        ws.append(success_row_report) 
#         wb.remove_sheet(wb.get_sheet_names()[0])
        wb.save(filename)   


""" Sometimes on older billboard charts, two songs are grouped together in one position. """
def deal_with_two_songs(filename, position, song_name, artist_name, last_week, weeks_on) :
    add_num_missd = 0
    
    split = song_name.find("/")
    song_name_1 = song_name[:split]
    song_name_2 = song_name[split+1:]
    
    lyrics_1 = extract_billboard_lyrics(song_name_1, artist_name)
    lyrics_2 = extract_billboard_lyrics(song_name_2, artist_name)
        
    row_1 = [position, song_name_1, artist_name, lyrics_1, last_week, weeks_on]
    row_2 = [position, song_name_2, artist_name, lyrics_2, last_week, weeks_on]
    
    if lyrics_1 == '' :
        add_num_missd += 1
        row_1 = [position, song_name_1, artist_name, 'COULD NOT FIND']
    if lyrics_2 == '' :
        add_num_missd += 1
        row_2 = [position, song_name_2, artist_name, 'COULD NOT FIND']

    return (row_1, row_2, add_num_missd)


"""  """
def go_again() :
    print(sep + "\nExtraction is complete. Would you like to extract more lyrics?")
    ans = input("\nEnter 'y' to continue, or 'n' or 'q' to quit: ").lower().strip()
    
    if ans == 'y' :
        script_start(sep + "Pick another chart.")
    elif ans == 'n' or ans == 'q' :
        print(sep, "Thanks for using the program!")
        sys.exit()
    else :
        print(sep, "You have made an invalid selection. Please try again.")
        return go_again()
      
                       
""" """        
def create_metrolyrics_url(song_name, artist_name):
    song = cleanup_song(song_name)
    song = (song.replace(" ", "-"))
         
    artist = cleanup_artist(artist_name)
    if "&" in artist :
        split = artist.find(" & ")
        artist = artist[:split]
        
    artist = (artist.replace(" ", "-"))

    url = "http://www.metrolyrics.com/" + song + "-lyrics-" + artist + ".html"
    if "lyrics-the-" in url : url = url.replace('lyrics-the-', 'lyrics-')
    if "-+-" in url : url = url.replace("-+-", "-")

    return url


"""  """
def extract_metrolyrics(url) :
    
    html = requestURL(url)
    if html == '' : return ''
        
    soup = BeautifulSoup(html, 'lxml')
    
    lyrics = ''             
    verse_tags = soup.select('.verse')
    if len(verse_tags) == 0 :
        print("\tCould not find: " + url)
        return ''
    
    for tag in verse_tags :
        lyrics = lyrics + tag.text +"\n"
    
    return lyrics


"""  """
def create_azlyrics_url(song_name, artist_name):

    song = cleanup_song(song_name)
    if " & " in song :
        song.replace(" & ", "")
    song = song.replace(" ", "")
    
    artist = cleanup_artist(artist_name)
    if "&" in artist :
        split = artist.find(" & ")
        artist = artist[:split]
    artist = artist.replace(" ", "") 

    azlyrics_url = "https://www.azlyrics.com/lyrics/" + artist + "/" + song + ".html"
#     if "$" in azlyrics_url : url = azlyrics_url.replace("$", "s")

    return azlyrics_url


"""  """
def extract_azlyrics(url) :
    
    html = requestURL(url)
    if html == '' : return ''
        
    soup = BeautifulSoup(html, 'lxml')
    
    # get lyrics
    div_tags = soup.findAll("div", {"class":None})
    lyrics = div_tags[1].text
    
    return lyrics


"""  """
def create_genius_url(song_name, artist_name):    
    song = cleanup_song(song_name)
    song = song.replace(" ", "-")
    
    artist = cleanup_artist(artist_name)
    artist = artist.replace(" ", "-")

    
    url = "https://genius.com/" + artist + "-" + song + "-lyrics"
    if "-+-" in url : url = url.replace("-+-", "and")
    if "$" in url : url = url.replace("$", "-")
    if "&" in url : url = url.replace("&", "and")

    return url


"""  """
def extract_genius(url) :
    html = requestURL(url)
    if html == '' : return ''
        
    soup = BeautifulSoup(html, 'lxml')
    
#     #get song name and artist 
#     title = soup.find("title").text[:-16] # slice " | Genius Lyrics" our of title
    
    # get lyrics
    lyrics = soup.select(".lyrics")[0].text
    
    return lyrics


""" """
def createFilename(chart_url, csv=False) :
    name_and_date = chart_url[34:]
    split = name_and_date.find('/')        
    
    name = name_and_date[:split]
    date = name_and_date[split+1:]
    
    if csv :
        return (name, date)

    return name + " (" + date + ")"
    
                  
        
""" Shows user message containing possible keyboard inputs """
def display_options() :
    print("\tls -" + " "*10 + "display all chart names")
    print("\tq  -" + " "*10 + "exit program")    


""""""
def cleanup_song(name) :
    punctuation = '!"\'()*+,.:;<=>?@[\\]^_`{|}~'
    table = str.maketrans(dict.fromkeys(punctuation))
    name = name.translate(table)
            
    return name.lower()


""" Edit the artist name to conform to url standards. 

    First, in the case of multiple artists, include only the first artist name. 
    Remove any names associated with words like 'Featuing' or 'With' or simply 
    added with commas. 
    
    Next, check for special cases of artist names that include punctuation, e.g. 
    'P!nk". In such cases, the punctuation maps to a letter: '!' becomes 'i'. 
    
    Finally, remove any
    other punctuation that appears in the name, but does not map to a letter. E.g.
    'Panic! at the Disco' becomes 'Panic at the Disco.' 
    
    @return artist name edited for url inclusion """
def cleanup_artist(name) :
    # check for cutoff symbols (for songs with multiple artists, only the first name is kept)
    # e.g. "Queen Featuring David Bowie" -> "Queen". 
    symbol_list = ["Featuring", ",", "Feat", "With", " X ", " x "]
    cutoff = len(name)
    for symbol in symbol_list :
        if symbol in name :
            cutoff = name.find(symbol)

    name = (name[:cutoff]).strip()

    # check for special cases of punctuation in artist's name
    punct_name_dict = {"P!nk":"Pink", "A$ap Rocky":"Asap Rocky"}
    for i, j in punct_name_dict.items() :
        name = name.replace(i, j)

    # remove punctuation
    punctuation = '!"\'()*+,./:;<=>?@[\\]^_`{|}~'
    table = str.maketrans(dict.fromkeys(punctuation))
    name = name.translate(table)
            
    return name.lower()


""" lname: list that contains strings representing billboard chart names
    
    displays a formatted 2-column numbered list of billboard chart names. """
def ls(lname) :
    assert(isinstance(lname, list))
    print("\n")
 
    # break list up into 2 columns using 2 iterators
    for a, b in zip(*[iter(enumerate(lname, start=1))]*2) :
        print(str(a[0]) + ": " + "{0:<55s}".format(a[1]) + 
              str(b[0]) + ": " + "{0:<50s}".format(b[1]) )
    
    print("\n")


""" """
def getEarliestDate(chart_name) :
    # Pick absurdly early date; Billboard.com will automatically round up to earliest chart
    url = "https://www.billboard.com" + chart_name + "/1900-01-01"
    html = requestURL(url)
    
    soup = BeautifulSoup(html, 'lxml')
    date = soup.find("time")['datetime']
    
    return date


""""""
def validate(date1, date2=None) :
    
    try :
        date1 = datetime.datetime.strptime(date1, '%Y-%m-%d')
    except ValueError :
        print("Invalid date entered: " + date1)
        return '#'
    
    # if only one date provided
    if date2 == None :
        date1 = resetDate(date1.date())
        
        return str(date1)
    
    # if both dates are provided
    else :
        try :
            date2 = datetime.datetime.strptime(date2, '%Y-%m-%d')
        except ValueError :
            print("Invalid date entered: " + date2)
            return '#'            
    
        date1 = resetDate(date1.date())
        date2 = resetDate(date2.date())

        if str(date1) == str(date2) : 
            return str(date1)
        
        if date2 < date1 :
            print("Error: second date earlier than the first.")
            return "#"
        
        # get dates for every chart before starting and ending date
        all_chart_dates = []
        while date1 <= date2 :
            all_chart_dates.append(str(date1))
            date1 += datetime.timedelta(days = 7)
        
        return all_chart_dates


""" date: a datetime object representing a date falling on a week when a given 
          billboard chart was published 
          
    billboard charts are """
def resetDate(date) :
    
    day_of_week = date.strftime("%A")
    
    switchover_date = datetime.date(1961, 12, 25)
    
    if date <= switchover_date :
        while day_of_week != "Monday" :
            date = date + datetime.timedelta(days = 1)
            day_of_week = date.strftime("%A")           
    else :
        while day_of_week != "Saturday" :
            date = date + datetime.timedelta(days = 1)
            day_of_week = date.strftime("%A")    
    
    return date


"""  """
def requestURL(url) :      
    hdrs = {'User-Agent':"Mozilla/5.0 (Macintosh; Intel Mac OS X 10_6_8) AppleWebKit/534.30 (KHTML, like Gecko) Chrome/12.0.742.112 Safari/534.30"}

    req = requests.get(url, headers = hdrs)
    if req.status_code == 200 :
        return req.text
    else :
        print("\tCould not find: " + url)
        return ''
    

if __name__=="__main__" :
    script_start(msg1)


    
